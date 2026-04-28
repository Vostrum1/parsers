import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from db_config import DB_CONFIG

def create_charts():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    wb = openpyxl.Workbook()

    # Дані для діаграм
    ws_data = wb.active
    ws_data.title = "Дані"

    # Середні ціни по продуктах
    cur.execute("""
        SELECT product, AVG(price_uah)::BIGINT
        FROM apple_ukraine WHERE price_uah>0
        GROUP BY product ORDER BY AVG(price_uah) DESC
    """)
    products = cur.fetchall()

    ws_data["A1"] = "Модель"
    ws_data["B1"] = "Середня ціна (грн)"
    ws_data["A1"].font = Font(bold=True)
    ws_data["B1"].font = Font(bold=True)
    for i, row in enumerate(products, 2):
        ws_data.cell(row=i, column=1, value=row[0])
        ws_data.cell(row=i, column=2, value=row[1])
    ws_data.column_dimensions["A"].width = 25
    ws_data.column_dimensions["B"].width = 20

    # Діаграма 1 - середні ціни по моделях
    ws_chart1 = wb.create_sheet("Ціни по моделях")
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.title = "Середні ціни Apple техніки в Україні"
    chart1.y_axis.title = "Ціна (грн)"
    chart1.x_axis.title = "Модель"
    chart1.width = 25
    chart1.height = 15
    data = Reference(ws_data, min_col=2, min_row=1, max_row=len(products)+1)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=len(products)+1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    ws_chart1.add_chart(chart1, "A1")

    # Дані магазинів
    ws_data2 = wb.create_sheet("Дані магазинів")
    cur.execute("""
        SELECT shop, COUNT(*), AVG(price_uah)::BIGINT
        FROM apple_ukraine WHERE price_uah>0
        GROUP BY shop ORDER BY COUNT(*) DESC
    """)
    shops = cur.fetchall()
    ws_data2["A1"] = "Магазин"
    ws_data2["B1"] = "Кількість товарів"
    ws_data2["C1"] = "Середня ціна"
    for i, row in enumerate(shops, 2):
        ws_data2.cell(row=i, column=1, value=row[0])
        ws_data2.cell(row=i, column=2, value=row[1])
        ws_data2.cell(row=i, column=3, value=row[2])

    # Діаграма 2 - кількість товарів по магазинах
    ws_chart2 = wb.create_sheet("Магазини")
    chart2 = BarChart()
    chart2.title = "Кількість товарів Apple по магазинах"
    chart2.y_axis.title = "Кількість"
    chart2.width = 20
    chart2.height = 12
    data2 = Reference(ws_data2, min_col=2, min_row=1, max_row=len(shops)+1)
    cats2 = Reference(ws_data2, min_col=1, min_row=2, max_row=len(shops)+1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws_chart2.add_chart(chart2, "A1")

    # Зведена таблиця найкращих цін
    ws_best = wb.create_sheet("Найкращі ціни")
    ws_best["A1"] = "Модель"
    ws_best["B1"] = "Магазин"
    ws_best["C1"] = "Ціна (грн)"
    ws_best["A1"].font = Font(bold=True)
    ws_best["B1"].font = Font(bold=True)
    ws_best["C1"].font = Font(bold=True)
    cur.execute("""
        SELECT DISTINCT ON (product) product, shop, price_uah
        FROM apple_ukraine WHERE price_uah>0
        ORDER BY product, price_uah ASC
    """)
    for i, row in enumerate(cur.fetchall(), 2):
        ws_best.cell(row=i, column=1, value=row[0])
        ws_best.cell(row=i, column=2, value=row[1])
        ws_best.cell(row=i, column=3, value=row[2])
        if i % 2 == 0:
            for col in range(1, 4):
                ws_best.cell(row=i, column=col).fill = PatternFill(fill_type="solid", fgColor="D6E4F0")
    ws_best.column_dimensions["A"].width = 25
    ws_best.column_dimensions["B"].width = 15
    ws_best.column_dimensions["C"].width = 15

    filename = "/home/vostrum/parsers/apple_charts.xlsx"
    wb.save(filename)
    print(f"Excel з діаграмами збережено: {filename}")
    cur.close()
    conn.close()

if __name__ == "__main__":
    create_charts()
