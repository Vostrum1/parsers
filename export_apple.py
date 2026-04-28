import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db_config import DB_CONFIG

def export():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    even_fill = PatternFill(fill_type="solid", fgColor="D6E4F0")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

    # Вкладка 1 - всі товари по ціні
    ws1 = wb.active
    ws1.title = "Всі товари"
    style_header(ws1, ["№","Магазин","Товар","Ціна (грн)","Посилання"])
    cur.execute("SELECT shop,product,price_uah,price_text,url FROM apple_ukraine WHERE price_uah>0 ORDER BY product,price_uah ASC")
    for i, row in enumerate(cur.fetchall(), 2):
        ws1.cell(row=i, column=1, value=i-1)
        for j, val in enumerate(row, 2):
            cell = ws1.cell(row=i, column=j, value=val)
            cell.border = thin
            if i % 2 == 0:
                cell.fill = even_fill
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 35
    ws1.column_dimensions["D"].width = 15
    ws1.column_dimensions["E"].width = 60

    # Вкладка 2 - порівняння цін по магазинах
    ws2 = wb.create_sheet("Порівняння магазинів")
    style_header(ws2, ["Магазин","Кількість товарів","Мін ціна","Макс ціна","Середня ціна"])
    cur.execute("""
        SELECT shop, COUNT(*), MIN(price_uah), MAX(price_uah), AVG(price_uah)::BIGINT
        FROM apple_ukraine WHERE price_uah>0
        GROUP BY shop ORDER BY COUNT(*) DESC
    """)
    for i, row in enumerate(cur.fetchall(), 2):
        for j, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=j, value=val)
            cell.border = thin
            if i % 2 == 0:
                cell.fill = even_fill
    for col in ["A","B","C","D","E"]:
        ws2.column_dimensions[col].width = 20

    # Вкладка 3 - найдешевший товар по кожній моделі
    ws3 = wb.create_sheet("Найкращі ціни")
    style_header(ws3, ["Товар","Магазин","Мін ціна","Посилання"])
    cur.execute("""
        SELECT DISTINCT ON (product) product, shop, price_uah, url
        FROM apple_ukraine WHERE price_uah>0
        ORDER BY product, price_uah ASC
    """)
    for i, row in enumerate(cur.fetchall(), 2):
        for j, val in enumerate(row, 1):
            cell = ws3.cell(row=i, column=j, value=val)
            cell.border = thin
            if i % 2 == 0:
                cell.fill = even_fill
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 60

    # Вкладка 4 - статистика
    ws4 = wb.create_sheet("Статистика")
    ws4["A1"] = "Аналіз цін Apple техніки в Україні"
    ws4["A1"].font = Font(bold=True, size=14, color="1F4E79")
    cur.execute("SELECT COUNT(*) FROM apple_ukraine WHERE price_uah>0")
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT shop) FROM apple_ukraine")
    shops = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT product) FROM apple_ukraine")
    products = cur.fetchone()[0]
    cur.execute("SELECT AVG(price_uah)::BIGINT FROM apple_ukraine WHERE price_uah>0")
    avg = cur.fetchone()[0]

    stats = [
        ("Всього записів", total),
        ("Магазинів", shops),
        ("Моделей товарів", products),
        ("Середня ціна", f"{avg:,} грн"),
    ]
    for i, (k, v) in enumerate(stats, 3):
        ws4.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws4.cell(row=i, column=2, value=v)
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 20

    filename = "/home/vostrum/parsers/apple_ukraine.xlsx"
    wb.save(filename)
    print(f"Excel збережено: {filename}")
    cur.close()
    conn.close()

if __name__ == "__main__":
    export()
